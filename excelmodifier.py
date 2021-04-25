import openpyxl
import requests
from bs4 import BeautifulSoup
import datetime
import os

class ExcelProcessor:
    #you may name the product sheet name here
    def __init__(self, filepath='EZDC.xlsx', product_sheet_name='Sheet1'):
        self.path = os.path.dirname(os.path.abspath(__file__))
        self.filename = filepath
        self.file = os.path.join(self.path, self.filename)
        self.wb = openpyxl.load_workbook(filename=self.file)
        self.product_sheet_name = product_sheet_name
        self.sheet_names = self.wb.sheetnames
        self.product_list = self.getProducts()
        self.session = requests.session()
                

    def getProducts(self):
        ws = self.wb[self.product_sheet_name]
        products = []
        for r in range(ws.max_row):
            #you may name the column number here
            products.append(ws.cell(row=r+1, column=4).value)
        # this line start to read the ASIN from row2, so if you want to start from row 3 or more, please do the addition (eg: start from row2 -> start from row3, [1:]->[2:])
        return products[1:]

    def updateWorkSheets(self):
        for product in self.product_list:
            product_asin = str(product)
            if product_asin not in self.sheet_names:
                if len(product_asin)>0:
                    print(f'New sheet created for {product_asin}!')
                    product_ws = self.wb.create_sheet(title=product_asin)
                    product_ws.append(('Date', 'ASIN', 'STARS', 'NUM_RATING', 'RANKING1','RANKING1_CAT','RANKING2', 'RANKING2_CAT'))
            else:
                product_ws = self.wb[product_asin]

            Date = datetime.datetime.today()

            try:
                ASIN, REVIEW, REVIEW_TEXT, rankings = self.getStarsReview(session=self.session, ASIN=product_asin)
                product_ws.append((Date, ASIN, REVIEW, REVIEW_TEXT, rankings[0][0], rankings[0][1], rankings[1][0], rankings[1][1]))
                print(f'Update Successful on {ASIN}.')
                print("DATE: "+str(Date.date()))
                print("ASIN: "+str(ASIN))
                print("STARS: "+str(REVIEW))
                print("RATING COUNTS: "+str(REVIEW_TEXT))
                print("RANKING1: "+str(rankings[0][0])+" in "+str(rankings[0][1]))
                print("RANKING2: "+str(rankings[1][0])+" in "+str(rankings[1][1]))
                print("----------------------------------------------------------")
            except:
                print(f'No page found for {product_asin}.')
                product_ws.append((Date, product_asin, 'NAN', 'NAN', 'NAN', 'NAN', 'NAN', 'NAN'))
        
        
        self.wb.save(self.file)
        print("Update finished. Bye!")

    def getStarsReview(self, session, ASIN):
        HEADERS = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0", "Accept-Encoding":"gzip, deflate", "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "DNT":"1","Connection":"close", "Upgrade-Insecure-Requests":"1"}
        url = 'https://www.amazon.com/dp/'+ASIN
        r = session.get(url, headers=HEADERS)
        s = BeautifulSoup(r.content, 'lxml')
        result = s.find_all('table', {'class':'prodDetTable'})
        ASIN = 'CHECKED EMPTY'
        REVIEW = 'CHECKED EMPTY'
        REVIEW_TEXT = 'CHECKED EMPTY'
        rankings = []
        for t in result:
            entries = t.find_all('th',{'class':'prodDetSectionEntry'})
            for ent in entries:
                if 'ASIN' in ent.text:
                    try:
                        ASIN = ent.nextSibling.nextSibling.string.replace('\n','')

                    except:
                        pass
                if 'Reviews' in ent.text:
                    try:
                        REVIEW = float(ent.nextSibling.nextSibling.find('span',{'id':'acrPopover'})['title'].replace(',','').replace(' out of 5 stars',''))
                        REVIEW_TEXT = int(ent.nextSibling.nextSibling.find('span',{'id':'acrCustomerReviewText'}).string.replace(',','').replace('\n','').replace(
                        'ratings','').replace(' ','').replace('rating',''))
                    except Exception as e:
                        print(e)
                if 'Rank' in ent.text:
                    try:
                        RANKING = ent.nextSibling.nextSibling
                        for r in RANKING.find_all('span'):
                            text = r.text
                            if '(' in text:
                                text = text[:text.index('(')]
                            texts = text.replace('#','').replace(r'\n','').replace('&amp;','&').split(' in ')
                            rank_num = int(texts[0].replace(',','').strip())
                            rank_cat = texts[1].strip()
                            rankings.append((rank_num, rank_cat))
                    except:
                        rankings = [('CHECKED EMPTY','CHECKED EMPTY'),('CHECKED EMPTY','CHECKED EMPTY')]
                 
        return(ASIN, REVIEW, REVIEW_TEXT, rankings[1:])

def Main():
    # you may make change to the filename here
    filename = 'EZDC.xlsx'
    EP = ExcelProcessor(filepath=filename)
    mtime = os.path.getmtime(EP.file)
    if datetime.datetime.today().date()!=datetime.datetime.utcfromtimestamp(mtime).date():
        EP.updateWorkSheets() 
    else:
        print(f'{filename} has been modified today, are you sure to update?(y/n)')
        ans = str(input())
        if ans == 'y' or ans == 'Y':
            EP.updateWorkSheets() 
        else:
            print('Bye~')

if __name__ =='__main__':
    Main()